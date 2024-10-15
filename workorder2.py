import openpyxl as excels
import cx_Oracle
import win32com.client as win32
from jinja2 import Template
import datetime
import re
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
import os
import shutil
import win32com.client

     
def flow():

        def has_special_chars(data):
            pattern = r"[^a-zA-Z0-9\s]"  
            return bool(re.search(pattern, data))

        def send_email(text_content,rendered_html, text_content2,recipient_email):
            outlook = win32.Dispatch('outlook.application')
            mail = outlook.CreateItem(0)
            mail.Subject = 'Health Check'
            mail.HTMLBody = text_content + "<br><br>" + rendered_html + "<br><br>" + text_content2 
            mail.To = recipient_email
            mail.Send()

        def updateExcel(row):

            for worksheet_row in worksheet.iter_rows(min_row=2): 
        
                if worksheet_row[3].value == row[3]:  # Assuming row[1] is the unique identifier
                    # Update the corresponding cell in column 3 (index 2) to "Tracker+Workflow AQ"
                    worksheet_row[2].value = "Tracker+Workflow AQ"  

        def PegaPush(row):
            print("Pega Push", row)
                                    
            tracker_query = f"""select Application, broker_contact_info, BROKER_CONTACT_PHONE,BROKER_CONTACT_EMAIL, BROKER_CODE,UMR, SUBMISSION_TYPE,
MARKET, CLASS_OF_BUSINESS, SLIP_TYPE, POLICY_TYPE, PROCESSING_REQUIRED, COUNT_OF_PREMIUM,COUNT_OF_AP,XIS_CONTACT_INFO,
Additional_information,PRESENTATIONDATE, WORKORDER_REF, GROUP_REFERENCE, NO_IN_GROUP, SIM_SIGNING_REQ,workorder_tag, TREATY_FDO_STATEMENT,
TRACKER_ID from repository.tblworkorder where workorder_ref='{row[22]}'"""
            
            t=[]
            for i in cur.execute(tracker_query):
                t.append(i) 

            for r in t:
                xml_code=f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<?xmlstylesheet type="text/xsl" href="/navigator/plugin/NativeRepositoryPlugin/getResource/nativeRepositoryPluginDojo/templates/WorkOrder.xslt"?>
<WorkOrder>
<TechAccountUUId></TechAccountUUId>
<Application>{r[0]}</Application>
<WorkOrderUniqueId>{r[21]}</WorkOrderUniqueId>
<UrgentReference></UrgentReference>
<BrokerContactName>{r[1]}</BrokerContactName>
<BrokerContactPhone>{r[2]}</BrokerContactPhone>
<BrokerContactEmail>{r[3]}</BrokerContactEmail>
<Broker>{r[4]}</Broker>
<UMR>{r[5]}</UMR>
<TypeOfSubmission>{r[6]}</TypeOfSubmission>
<Market>{r[7]}</Market>
<ClassOfBusiness>{r[8]}</ClassOfBusiness>
<SlipType>{r[9]}</SlipType>
<TypeOfPolicy>{r[10]}</TypeOfPolicy>
<ProcessingRequired>{r[11]}</ProcessingRequired>
<PremiumFDOLPANS>{r[12]}</PremiumFDOLPANS>
<APRPLPANS>{r[13]}</APRPLPANS>
<XISContactName>{r[14]}</XISContactName>
<AdditionalDetails>{r[15]}</AdditionalDetails>
<PresentationDate>{r[16]}</PresentationDate>
<WorkOrderReference>{r[17]}</WorkOrderReference>
<GroupReference>{r[18]}</GroupReference>
<NumberInGroup>{r[19]}</NumberInGroup>
<UCR></UCR>
<TR></TR>
<SimSigningRequired>{r[20]}</SimSigningRequired>
<TreatyFdoStatement>{r[22]}</TreatyFdoStatement>
<XisScanned></XisScanned>
<documents>
<Document><Type></Type><ID></ID><Reference></Reference><Version></Version><Description></Description></Document>
<Document><Type></Type><ID></ID><Reference></Reference><Version></Version><Description></Description></Document>
<Document><Type></Type><ID></ID><Reference></Reference><Version></Version><Description></Description></Document>
<Document><Type></Type><ID></ID><Reference></Reference><Version></Version><Description></Description></Document>
<Document><Type></Type><ID></ID><Reference></Reference><Version></Version><Description></Description></Document></documents>
<trackerBarcode>{r[23]}</trackerBarcode>
<logisticsInView>
http://insuranceportal.xchanging.com/doctrack/DataCapture.jsp?barCode=REP02891246&amp;sourcepage=XisWorkNotPrintedTableTag</logisticsInView>
<logisticsOutView>
http://insuranceportal.xchanging.com/doctrack/Departures?barCode=REP02891246</logisticsOutView>
<techniciansView>
http://insuranceportal.xchanging.com/doctrack/TransactionStatusArrivals?barCode=REP02891246</techniciansView>
<workPackageView>
https://repository.xchanging.com/web/?desktop=ui&amp;feature=default&amp;wpr=BHCYIGE</workPackageView>
<PegaWO></PegaWO></WorkOrder>"""

                url = "http://amqadmin:CBwY=y4(c@+@basimrprdappz16.xchanginghosting.com:8080/admin/queues.jsp"

                driver = webdriver.Chrome()

                driver.get(url)

                anchor_element = driver.find_element(By.XPATH, "//*[@id='queues']/tbody/tr[37]/td[1]/a")
                anchor_element.click()
                time.sleep(5)

                send = driver.find_element(By.XPATH, '//*[@id="site-breadcrumbs"]/a[8]')
                send.click()
                time.sleep(5)

                textarea = driver.find_element(By.XPATH, '//*[@id="headers"]/tbody/tr[11]/td/textarea')
                textarea.send_keys(xml_code)
                time.sleep(2)

                sendxml = driver.find_element(By.XPATH, '//*[@id="headers"]/tbody/tr[9]/td/input[1]')
                sendxml.click()
                time.sleep(2)

                driver.quit()


            #update excel status to OK
            for worksheetrow in worksheet.iter_rows(min_row=1):
                for cell in worksheetrow:
                    if cell.value == row[22]:
                        worksheetrow[2].value = "OK" 
                    
        worksbook = excels.load_workbook(source_file)
        time.sleep(3)
        worksheet = worksbook.active

        tracker_Worflow_Email=[]

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if row[2] =='Tracker+Workflow AQ+Email':
                tracker_Worflow_Email.append(row)

        missing_status_rows = []

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if (row[2] == 'missing'):  #or (row[2] =='Tracker+Workflow AQ'))
                missing_status_rows.append(row)

        # workOrderRef Index 
        index_to_extract = 22 

        # Here We get WorkorderRefernce in a list
        WORKORDER_REF = [t[index_to_extract] for t in missing_status_rows]

        # Now I want to fetch the status of these workorder ref

        # Construct the IN clause(The IN clause will generate the required SQL query)
        in_clause = ", ".join(["'" + value + "'" for value in WORKORDER_REF])

        sql_query = f"SELECT work_order_status FROM repository.tblworkorder WHERE workorder_ref IN ({in_clause})"

        connection = cx_Oracle.connect('Pankaj/Dxc1234@10.5.2.43:1521/imrprd.xchanging.com')

        cur = connection.cursor()

        workOrder_Status = []  #initails a list so it will store workorder status 

        # Execute the SQL query and store the values
        for qw in cur.execute(sql_query):
            workOrder_Status.append(qw)   

        # Now we have workorder status, So verify it if it is in new state or not
        if any(status[0] != 'CAN' for status in workOrder_Status):
            print("There is at least one work order with a status other than 'NEW'.")
        #if all are in new state check the three columns    
        else:

            #Firstly what I do here, I distribute missing rows based on client
            email_dict = {}
            # Iterate through the tuples and categorize them based on email initials
            for row in missing_status_rows:
                email = row[8]  # Index 8 contains the email
                if email not in email_dict: 
                    email_dict[email] = []        #here if client not found previously it will create array for that client
                email_dict[email].append(row)     #here it will add rows of a particular client in a array

            # number_of_clients = len(email_dict.keys())
            for email in email_dict:
                client_rows = email_dict[email]
                for row in client_rows:
                    
                    EmailDATE = row[16]

                    if row[6] is not None and has_special_chars(row[6]):
                        recipient_email = "ankit.raghav@dxc.com"

                        text_content=f"""Hi, <br><br>Below is the Work package that you have submitted on {row[16]}, which has failed to reach some of the required systems at our end.<br><br>
                        The details are as follows"""  

                        text_content2="""Please note that there should not be any junk characters in additional information. Also, we are cancelling these work package at our end please make a resubmission. If you have any further concerns please raise a call with the Xchanging Service Centre.<br><br><br>
                        Thanks and Regards<br>
                        Ankit<br>
                        DXC Technology<br>
                        """

                        html_content="""<!DOCTYPE html>
                        <html lang="en">
                        <head>
                        <meta charset="UTF-8">
                        <meta name="viewport" content="width=device-width, initial-scale=1.0">
                        <title>Table Example</title>
                        <style>
                            table {
                                border-collapse: collapse;
                                width: 50%;
                            }
                            th, td {
                                border: 1px solid black;
                                padding: 8px;
                                text-align: center;
                            }
                            th {
                                background-color: #f2f2f2;
                            }
                            td[colspan="3"] {
                                background-color: #c0c0c0;
                                font-weight: bold;
                            }
                        </style>
                        </head>
                        <body>
                        <table>
                        <thead>
                            <tr>
                            <th>UMR</th>
                            <th>WORKORDER_TAG</th> 
                            <th>Additional Information</th>
                            <th>Reason</th>
                            <th>WORKORDER_REF</th>
                            </tr>
                        </thead>
                        <tbody>
                            {% for row in client_rows %}
                            <tr>
                            <td>{{ row[0] }}</td>
                            <td>{{ row[1] }}</td>
                            <td>{{ row[6] }}</td>
                            <td>Junk Characters</td>
                            <td>{{ row[22] }}</td>
                            </tr>
                            {% endfor %}
                        </tbody>
                        </table>

                        </body>
                        </html>
                        """

                        template = Template(html_content)

                        # Render the template with client_rows data
                        rendered_html = template.render(client_rows=client_rows)

                        send_email(text_content,rendered_html,text_content2,recipient_email)
                        #to update the excel from missing to tracker+workflowAQ
                        updateExcel(row)
                    
    # CASE 5   #Same thing we are doing in case1 If his case exist so it have check before case1
    # I think this case is not exist anymore
                    elif row[3] is None and row[5] is None and row[6] is None:  
                            PegaPush(row)
                            
                    #GroupRef is present
                    elif row[3] is not None:   

                        # Gourp Id is not none so we have to check they all are same or not
                        same_value = all(row2[3] == client_rows[0][3] for row2 in client_rows)
                        if same_value:
                        
                            desired_date_format = row[4] 

                            GrpRef_query = f"""SELECT UMR, Created_Date, Workorder_Ref, NO_IN_GROUP, WORKORDER_TAG, BROKER_CODE, GROUP_REFERENCE, WORK_ORDER_STATUS FROM repository.tblworkorder WHERE Group_Reference = '{row[3]}' AND created_date = TO_DATE('{desired_date_format}', 'DD-MM-YYYY')"""
                            
                            table=[]
                            for i in cur.execute(GrpRef_query):
                                table.append(i)   
                            # Now table have all the rows that have above query data

                            if any(status[7] != 'CAN' for status in table):
                                print("There is at least one work order with a status other than 'NEW' while checking for casee2")                    
                            # check status, we already did but still for confirmation
                            else:
                                def check_column(table):
                                    reference_value = table[0][5]
                                    different_value = None
                                    
                                    for row in table:
                                        if row[5] != reference_value:
                                            different_value = row[5]
                                            break
                                            
                                    return reference_value, different_value
                                
                                reference_value, different_value = check_column(table)
                                #this will store same brokercode as well as different broker code if there is any
                                
                                #CASE 2 
                                noInGroup=table[0][3]
                                submissions=len(table)
                                if different_value is None:
                                    
                                    if(submissions != noInGroup):
                                    
                                        recipient_email = "ankit.raghav@dxc.com"

                                        text_content=f"""Hi, <br><br>Below is the Work package that you have submitted on {EmailDATE}, which has failed to reach some of the required systems at our end.<br><br>
                                        The details are as follows"""  
                                        
                                        
                                        text_content2=f"""Please not that these submissions has been failed because you have entered no in group as {noInGroup} whereas {submissions} has been made.<br><br>
                                        Hence, we would advise you make these submission again and we will cancel the work packages at our end. If you have any further concerns please raise a call with the Xchanging Service Centre.<br><br><br>
                                        Thanks and Regards<br>
                                        Ankit<br>
                                        DXC Technology<br>
                                        """

                                        html_content="""<!DOCTYPE html>
                                        <html lang="en">
                                        <head>
                                        <meta charset="UTF-8">
                                        <meta name="viewport" content="width=device-width, initial-scale=1.0">
                                        <title>Table Example</title>
                                        <style>
                                            table {
                                                border-collapse: collapse;
                                                width: 50%;
                                            }
                                            th, td {
                                                border: 1px solid black;
                                                padding: 8px;
                                                text-align: center;
                                            }
                                            th {
                                                background-color: #f2f2f2;
                                            }
                                            td[colspan="3"] {
                                                background-color: #c0c0c0;
                                                font-weight: bold;
                                            }
                                        </style>
                                        </head>
                                        <body>
                                        <table>
                                        <thead>
                                            <tr>
                                            <th>UMR</th>
                                            <th>WORKORDER_TAG</th>
                                            <th>WORKORDER_REF</th>
                                            <th>Group_REF</th>
                                            <th>No_in_Group</th>
                                            </tr>
                                        </thead>
                                        <tbody>
                                            
                                            {% for row in client_rows%}
                                            <tr>
                                            <td>{{ row[0] }}</td>
                                            <td>{{ row[1] }}</td>
                                            <td>{{ row[22] }}</td>
                                            <td>{{ row[3] }}</td>
                                            <td>{{ noInGroup }}</td>
                                            </tr>
                                            {% endfor %}  
                                            
                                        </tbody>
                                        </table>

                                        </body>
                                        </html>
                                        """

                                        template = Template(html_content)

                                        # Render the template with client_rows data
                                        rendered_html = template.render(client_rows=client_rows,noInGroup=noInGroup)

                                        send_email(text_content,rendered_html,text_content2,recipient_email)

                                #Case3
                                elif different_value is not None:

                                    recipient_email = "ankit.raghav@dxc.com"

                                    text_content=f"""Hi, <br><br>Below is the Work package that you have submitted on {row[16]}, which has failed to reach some of the required systems at our end.<br><br>
                                    The details are as follows"""  

                                    text_content2="""Please note that these submissions have failed because you have made submission with different broker codes.<br><br>
                                    Hence, we would advise you make these submission again with same broker code and with correct number in group. The current work packages will be cancelled by the system at our end. If you have any further concerns please raise a call with the Xchanging Service Centre.<br><br><br>
                                    Thanks and Regards<br>
                                    Ankit<br>
                                    DXC Technology<br>
                                    """

                                    html_content="""<!DOCTYPE html>
                                    <html lang="en">
                                    <head>
                                    <meta charset="UTF-8">
                                    <meta name="viewport" content="width=device-width, initial-scale=1.0">
                                    <title>Table Example</title>
                                    <style>
                                        table {
                                            border-collapse: collapse;
                                            width: 50%;
                                        }
                                        th, td {
                                            border: 1px solid black;
                                            padding: 8px;
                                            text-align: center;
                                        }
                                        th {
                                            background-color: #f2f2f2;
                                        }
                                        td[colspan="3"] {
                                            background-color: #c0c0c0;
                                            font-weight: bold;
                                        }
                                    </style>
                                    </head>
                                    <body>
                                    <table>
                                    <thead>
                                        <tr>
                                        <th>UMR</th>
                                        <th>WORKORDER_TAG</th>
                                        <th>WORKORDER_REF</th>
                                        <th>Group Reference</th>
                                        <th>No_In_Group</th>
                                        <th>Broker Number</th>
                                        </tr>
                                    </thead>
                                    <tbody>
                                        {% set is_first_row = True %}
                                        {% for row in client_rows %}
                                        <tr>
                                        <td>{{ row[0] }}</td>
                                        <td>{{ row[1] }}</td>
                                        <td>{{ row[22] }}</td>
                                        {% if is_first_row %}
                                        <td>{{ row[3]  }}</td>
                                        <td>{{ noInGroup }}</td>
                                        <td>{{ reference_value,different_value }}</td>
                                        {% else %}
                                        <td></td>
                                        <td></td>
                                        <td></td>
                                        {% endif %}
                                        {% set is_first_row = False %}
                                        </tr>
                                        {% endfor %}
                                    </tbody>
                                    </table>

                                    </body>
                                    </html>
                                    """

                                    template = Template(html_content)

                                    rendered_html = template.render(client_rows=client_rows,noInGroup=noInGroup,reference_value=reference_value, different_value=different_value)

                                    send_email(text_content,rendered_html,text_content2,recipient_email)

                                else:#(wog case)
                                    pass


                                #to update the excel from missing to tracker+workflowAQ
                                updateExcel(row)

                        else:
                            print("Some rows have different GroupRef")
                        break


            case1workorder=[]
            for row in missing_status_rows:
                if row[3] is None and row[5] is None:
                    case1workorder.append(row)

            driver = webdriver.Chrome()
            driver.get('https://repository.xchanging.com/web/')
            driver.maximize_window()

            time.sleep(20)

            Account_input = driver.find_element(By.ID , "nativeRepositoryPluginDojo_IMRLayout_0_LoginPane_XCusername")
            Account_input.send_keys("LPCI")

            username_input = driver.find_element("name","nativeRepositoryPluginDojo_IMRLayout_0_LoginPane_XCusername2")
            username_input.send_keys("IND16Q")

            time.sleep(2)

            password_input = driver.find_element("name", "nativeRepositoryPluginDojo_IMRLayout_0_LoginPane_password")
            password_input.send_keys("Partnership@452")
            password_input.send_keys(Keys.ENTER)

            time.sleep(15)

            workPackage_click = driver.find_element("id", "dijit__TreeNode_3_label")
            workPackage_click.click()
            time.sleep(2)
            
            Technical_Issue_Workpackage=[]

            for value in case1workorder:

                Search_Value= driver.find_element("id", "iMRSearchTemplatePluginDojo_SearchForm_1_ecm.widget.SearchCriterian_0")
                Search_Value.clear()
                Search_Value.send_keys(value[22])
                time.sleep(2)

                Search_click = driver.find_element("id", "dijit_form_Button_27_label")
                Search_click.click()
                time.sleep(12)


                verify_Contact= driver.find_element("id", "contractDetail_titleBarNode")
                if(verify_Contact.text=='Contract Details'):
                    pass
                else:
                    Technical_Issue_Workpackage.append(value)
                    
                time.sleep(5)
                search_icon=driver.find_element(By.CSS_SELECTOR, ".dijitReset.dijitInline.dijitIcon.iconNode.IMRSearchTemplatePluginLaunchIcon")
                search_icon.click()
                time.sleep(5)

                Search_criteria=driver.find_element(By.XPATH,'//*[@id="dijit_layout_ContentPane_33"]/div[1]/div/table/tbody/tr/td[1]/div')
                Search_criteria.click()
                time.sleep(2)

            PushWorkpackages = list(set(case1workorder) - set(Technical_Issue_Workpackage))

            for not_shown_package_row in Technical_Issue_Workpackage:
                recipient_email = "ankit.raghav@dxc.com"

                text_content=f"""Hi, <br><br>Below is the Work package that you have submitted on {not_shown_package_row[16]}, which has failed to reach some of the required systems at our end.<br><br>
                The details are as follows"""  

                text_content2="""Due to some technical issue above submissions have failed to reach to some of our systems.<br><br>
                Hence, we would advise you make these submission again and we will cancel the work packages at our end. If you have any further concerns please raise a call with the Xchanging Service Centre.<br><br><br>
                Thanks and Regards<br>
                Ankit<br>
                DXC Technology<br>
                """

                html_content="""<!DOCTYPE html>
                <html lang="en">
                <head>
                <meta charset="UTF-8">
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                <title>Table Example</title>
                <style>
                    table {
                        border-collapse: collapse;
                        width: 50%;
                    }
                    th, td {
                        border: 1px solid black;
                        padding: 8px;
                        text-align: center;
                    }
                    th {
                        background-color: #f2f2f2;
                    }
                    td[colspan="3"] {
                        background-color: #c0c0c0;
                        font-weight: bold;
                    }
                </style>
                </head>
                <body>
                <table>
                <thead>
                    <tr>
                    <th>UMR</th>
                    <th>WORKORDER_TAG</th>
                    <th>WORKORDER_REF</th>
                    </tr>
                </thead>
                <tbody>
                    {% for not_shown_package_row in NotShown_Workpackage %}
                    <tr>
                    <td>{{ not_shown_package_row[0] }}</td>
                    <td>{{ not_shown_package_row[1] }}</td>
                    <td>{{ not_shown_package_row[22] }}</td>
                    </tr>
                    {% endfor %}
                </tbody>
                </table>

                </body>
                </html>
                """

                template = Template(html_content)

                rendered_html = template.render(client_rows=client_rows)

                send_email(text_content,rendered_html,text_content2,recipient_email)

                #to update the excel from missing to tracker+workflowAQ
                updateExcel(row)   

            for push_work_package_row in PushWorkpackages:
                PegaPush(push_work_package_row)
                #TrackerPush(push_work_package_row)
    
        cur.close()

        connection.close()

        filenameDate = os.path.splitext(row[4])[0].replace("\\", "_").replace("/", "_") 

        worksbook.save(f'UpdatedImrExcel{filenameDate}.xlsx')


def process_excel_file(source_file, destination_folder):
    

    flow()
   
    # shutil.move(source_file, destination_folder)
    os.remove(source_file)


# Folder to save attachments
save_folder = r"C:\Users\araghav6\OneDrive - DXC Production\Desktop\workorderfile"

# Create the save folder if it doesn't exist
if not os.path.exists(save_folder):
    os.makedirs(save_folder)

outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")

# Access the inbox (folder index 6 is the inbox)
inbox = outlook.GetDefaultFolder(6)

# Get all items in the inbox
messages = inbox.Items

# Function to save attachments from a message
def save_attachments(message, save_folder):
    attachments = message.Attachments
    for attachment in attachments:
        if attachment.FileName.lower().endswith(".xlsx"):
            attachment.SaveAsFile(os.path.join(save_folder, attachment.FileName))
            

# Process each message
for message in messages:
    
    try:
        subject = message.Subject
        if "Work Order Report" in subject:
            # print(f"Processing email with subject: {subject}")
            save_attachments(message, save_folder)
    except Exception as e:
        print(f"Error processing message: {e}")


#I used save and source folder two different variable but they have same value
source_folder = r"C:\Users\araghav6\OneDrive - DXC Production\Desktop\workorderfile" 
destination_folder = r"C:\Users\araghav6\Downloads\workorder"  

# Loop until there are no more Excel files in the source folder
while True:
    excel_files = [f for f in os.listdir(source_folder) if f.endswith(".xlsx")]

    if not excel_files:
        print("No more Excel files found in the source folder.")
        break
    

    for file in excel_files:
        source_file = os.path.join(source_folder, file)
        process_excel_file(source_file, destination_folder)  







